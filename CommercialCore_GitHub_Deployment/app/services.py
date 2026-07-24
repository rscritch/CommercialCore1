from __future__ import annotations
from datetime import date, timedelta, datetime
from decimal import Decimal
import csv, io, hashlib, json
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy import select, func
from sqlalchemy.orm import Session
from .models import (
    Exposure, ReportingEntry, Projection, ReviewItem, AuditEvent, ImportFile
)
from .config import UPLOAD_DIR
from .intelligence import build_intelligence_score, SCORING_VERSION

def audit(db: Session, actor_id: int | None, entity_type: str, entity_id: object, action: str, details: dict | str):
    event = AuditEvent(
        actor_id=actor_id,
        entity_type=entity_type,
        entity_id=str(entity_id),
        action=action,
        details=json.dumps(details, default=str) if isinstance(details, dict) else str(details),
    )
    db.add(event)

def overlap_exists(db: Session, exposure_id: int, start: date, end: date) -> bool:
    stmt = select(ReportingEntry.id).where(
        ReportingEntry.exposure_id == exposure_id,
        ReportingEntry.accepted.is_(True),
        ReportingEntry.period_start <= end,
        ReportingEntry.period_end >= start,
    )
    return db.execute(stmt).first() is not None

def accepted_entries(db: Session, exposure_id: int):
    stmt = select(ReportingEntry).where(
        ReportingEntry.exposure_id == exposure_id,
        ReportingEntry.accepted.is_(True),
    ).order_by(ReportingEntry.period_start)
    return list(db.scalars(stmt))

def policy_days(exposure: Exposure) -> int:
    return max((exposure.policy.expiration_date - exposure.policy.effective_date).days + 1, 1)

def elapsed_days(exposure: Exposure, as_of: date) -> int:
    end = min(as_of, exposure.policy.expiration_date)
    if end < exposure.policy.effective_date:
        return 0
    return (end - exposure.policy.effective_date).days + 1

def calculate_projection(db: Session, exposure: Exposure, as_of: date | None = None) -> Projection:
    as_of = as_of or date.today()
    entries = accepted_entries(db, exposure.id)
    current = [
        e for e in entries
        if e.period_start >= exposure.policy.effective_date and e.period_end <= min(as_of, exposure.policy.expiration_date)
    ]
    current_total = sum(float(e.actual_value) for e in current)
    total_days = policy_days(exposure)
    elapsed = max(elapsed_days(exposure, as_of), 1)
    history_by_year = {}
    for e in entries:
        if e.period_end < exposure.policy.effective_date:
            history_by_year.setdefault(e.period_end.year, []).append(e)
    complete_years = []
    for year, rows in history_by_year.items():
        if len(rows) >= {"weekly": 45, "monthly": 10, "quarterly": 3}.get(exposure.cadence, 10):
            complete_years.append(year)

    method = "straight_line"
    explanation = ""
    projected = current_total / elapsed * total_days if current_total else float(exposure.recorded_estimate)

    if len(complete_years) >= 2 and current:
        ratios = []
        for year in sorted(complete_years)[-3:]:
            rows = history_by_year[year]
            total = sum(float(r.actual_value) for r in rows)
            cutoff_index = min(len(current), len(rows))
            pace = sum(float(r.actual_value) for r in rows[:cutoff_index])
            if pace > 0:
                ratios.append(total / pace)
        if ratios:
            factor = sum(ratios) / len(ratios)
            projected = current_total * factor
            method = "seasonal"
            explanation = f"Seasonal projection used {len(ratios)} comparable prior years and current accepted total {current_total:,.2f}."
    elif len(complete_years) == 1 and current:
        year = complete_years[0]
        rows = history_by_year[year]
        total = sum(float(r.actual_value) for r in rows)
        cutoff_index = min(len(current), len(rows))
        pace = sum(float(r.actual_value) for r in rows[:cutoff_index])
        if pace > 0:
            projected = current_total * (total / pace)
            method = "prior_year_pace"
            explanation = f"Prior-year pace projection used {year} as the comparable year."
    if not explanation:
        explanation = f"Straight-line projection annualized {current_total:,.2f} across {elapsed} elapsed policy days."

    estimate = float(exposure.recorded_estimate)
    variance = ((projected - estimate) / estimate * 100) if estimate else 0.0
    expected = {"weekly": 52, "monthly": 12, "quarterly": 4}.get(exposure.cadence, 12)
    elapsed_share = min(elapsed / total_days, 1)
    expected_to_date = max(round(expected * elapsed_share), 1)
    latest_period_end = max((e.period_end for e in current), default=None)
    days_late = max((as_of - latest_period_end).days - exposure.reporting_due_days, 0) if latest_period_end else 0
    business = exposure.policy.business
    open_reviews = [r for r in business.reviews if r.status != "closed" and (r.exposure_id in (None, exposure.id))]
    high_reviews = [r for r in open_reviews if r.priority == "high"]
    days_to_renewal = (exposure.policy.expiration_date - as_of).days if exposure.policy.expiration_date else None
    intelligence = build_intelligence_score(
        variance_percent=variance,
        received=len(current),
        expected_to_date=expected_to_date,
        days_late=days_late,
        complete_years=len(complete_years),
        days_to_renewal=days_to_renewal,
        open_reviews=len(open_reviews),
        high_reviews=len(high_reviews),
        has_estimate=estimate > 0,
        has_current_reporting=bool(current),
        projection_method=method,
    )
    accuracy = intelligence["accuracy_score"]
    confidence = intelligence["confidence_score"]
    core_index = intelligence["core_index"]

    projection = Projection(
        exposure_id=exposure.id,
        as_of_date=as_of,
        method=method,
        projected_total=round(projected, 2),
        variance_percent=round(variance, 3),
        accuracy_score=accuracy,
        confidence_score=confidence,
        core_index=core_index,
        explanation=explanation,
        scoring_version=SCORING_VERSION,
        score_details=json.dumps(intelligence),
    )
    db.add(projection)
    db.flush()
    evaluate_rules(db, exposure, projection)
    return projection

def evaluate_rules(db: Session, exposure: Exposure, projection: Projection):
    variance = abs(float(projection.variance_percent))
    desired = None
    if variance >= 10:
        desired = ("material_variance", "high", "Review suggested: material exposure variance")
    elif variance >= 5:
        desired = ("watch_variance", "medium", "Exposure variance watch")
    elif projection.confidence_score < 70:
        desired = ("low_confidence", "medium", "Reporting confidence below threshold")
    if not desired:
        return
    rule_code, priority, title = desired
    existing = db.scalar(select(ReviewItem).where(
        ReviewItem.exposure_id == exposure.id,
        ReviewItem.rule_code == rule_code,
        ReviewItem.status.in_(["open", "in_progress"]),
    ))
    if existing:
        existing.evidence = f"Projection {float(projection.projected_total):,.2f}; variance {float(projection.variance_percent):+.1f}%; confidence {projection.confidence_score}."
        return
    item = ReviewItem(
        business_id=exposure.policy.business_id,
        exposure_id=exposure.id,
        rule_code=rule_code,
        priority=priority,
        title=title,
        evidence=f"Projection {float(projection.projected_total):,.2f}; variance {float(projection.variance_percent):+.1f}%; confidence {projection.confidence_score}.",
        assigned_to=exposure.policy.business.producer_id,
    )
    db.add(item)


def parse_import(filename: str, content: bytes):
    suffix = Path(filename).suffix.lower()
    rows = []
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append(row)
    elif suffix in (".xlsx", ".xlsm"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        if not data:
            return []
        headers = [str(v).strip() if v is not None else "" for v in data[0]]
        for values in data[1:]:
            rows.append(dict(zip(headers, values)))
    else:
        raise ValueError("Only CSV and XLSX files are supported.")
    return rows

def normalize_import_row(row: dict):
    lower = {str(k).strip().lower(): v for k, v in row.items()}
    def get(*names):
        for n in names:
            if n in lower:
                return lower[n]
        return None
    start = get("period_start", "start", "start_date")
    end = get("period_end", "end", "end_date")
    value = get("actual_value", "amount", "payroll", "value")
    if start is None or end is None or value is None:
        raise ValueError("Required columns: period_start, period_end, actual_value")
    if hasattr(start, "date"):
        start = start.date()
    elif not isinstance(start, date):
        start = date.fromisoformat(str(start).strip())
    if hasattr(end, "date"):
        end = end.date()
    elif not isinstance(end, date):
        end = date.fromisoformat(str(end).strip())
    value = float(str(value).replace(",", "").replace("$", ""))
    if value < 0 or end < start:
        raise ValueError("Invalid date range or negative value.")
    return start, end, value

def store_import(db: Session, exposure: Exposure, user_id: int, filename: str, content: bytes):
    digest = hashlib.sha256(content).hexdigest()
    if db.scalar(select(ImportFile).where(ImportFile.sha256 == digest)):
        raise ValueError("This exact file has already been uploaded.")
    stored_name = f"{digest[:16]}_{Path(filename).name}"
    (UPLOAD_DIR / stored_name).write_bytes(content)
    record = ImportFile(
        exposure_id=exposure.id,
        original_name=filename,
        stored_name=stored_name,
        sha256=digest,
        status="processing",
        uploaded_by=user_id,
    )
    db.add(record)
    db.flush()
    rows = parse_import(filename, content)
    created = 0
    errors = []
    for idx, row in enumerate(rows, start=2):
        try:
            start, end, value = normalize_import_row(row)
            if overlap_exists(db, exposure.id, start, end):
                raise ValueError("Overlaps an existing accepted reporting period.")
            db.add(ReportingEntry(
                exposure_id=exposure.id,
                period_start=start,
                period_end=end,
                actual_value=value,
                source="upload",
                source_file=filename,
                created_by=user_id,
            ))
            db.flush()
            created += 1
        except Exception as exc:
            errors.append(f"Row {idx}: {exc}")
    record.row_count = created
    record.status = "accepted" if not errors else ("accepted_with_warnings" if created else "rejected")
    record.error_text = "\n".join(errors) if errors else None
    return record
